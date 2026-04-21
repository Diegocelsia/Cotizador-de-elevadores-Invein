import re
from openpyxl import load_workbook
from database import SessionLocal
from models import SheetPrice, BomTemplate

def clean_currency(value):
    """Limpia simbolos de moneda y formatos de texto a numero float"""
    if value is None or value == "" or value == "-":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    
    clean_str = re.sub(r'[^\d,.-]', '', str(value)).strip()
    if "." in clean_str and "," in clean_str:
        clean_str = clean_str.replace(".", "").replace(",", ".")
    elif "." in clean_str and len(clean_str.split(".")[-1]) == 3:
        clean_str = clean_str.replace(".", "")
        
    try:
        return float(clean_str)
    except ValueError:
        return 0.0


def is_lamina(material: str) -> bool:
    texto = (material or "").lower().replace("á", "a")
    return "lamina" in texto or re.search(r"\blam\b", texto) is not None


def extract_calibre_from_text(texto: str):
    normalizado = (texto or "").lower().replace("á", "a")
    match = re.search(r"cal\.?\s*(\d{1,2})", normalizado)
    if match:
        return int(match.group(1))
    return None


def infer_lamina_calibre(material: str) -> int:
    """Infiere un calibre de lámina para la UI de selección.

    El archivo Material.xlsx mezcla lámina genérica con descripciones de espesor.
    Para el selector de la aplicación usamos un calibre base válido.
    """
    texto = (material or "").lower()

    for calibre in (12, 14, 16, 18):
        if f"cal. {calibre}" in texto or f"cal {calibre}" in texto:
            return calibre

    # Si el material es lámina pero no trae un calibre estándar del catálogo,
    # usamos 12 como base de trabajo para que el usuario pueda ajustar.
    return 12


def read_xlsx_rows(path: str):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        item = {}
        for idx, header in enumerate(headers):
            if header:
                item[header] = row[idx] if idx < len(row) else None
        data.append(item)
    return data

def importar_datos():
    db = SessionLocal()
    try:
        print("[INFO] Iniciando importacion con limpieza de datos...")

        # 1. Cargar Lamina.xlsx
        rows_lamina = read_xlsx_rows("data/Lamina.xlsx")
        db.query(SheetPrice).delete()
        
        for row in rows_lamina:
            tipo = str(row.get('Tipo', '')).strip().upper()
            material_desc = str(row.get('Material', ''))
            calibre = extract_calibre_from_text(material_desc) or 12
            
            nueva = SheetPrice(
                material=tipo,
                calibre=calibre,
                peso_hoja=clean_currency(row.get('Peso')),
                valor_unitario=clean_currency(row.get('Valor Unit'))
            )
            db.add(nueva)
        print("[OK] Precios de lamina cargados con calibres normalizados.")

        # 2. Cargar Material.xlsx
        rows_material = read_xlsx_rows("data/Material.xlsx")
        db.query(BomTemplate).delete()
        for row in rows_material:
            material_text = str(row.get('Material', ''))
            seleccionar_lamina = str(row.get('Seleccionar lamina', 'No')).strip().lower() == 'si'

            calibre = None
            if is_lamina(material_text):
                calibre = extract_calibre_from_text(material_text) or infer_lamina_calibre(material_text)
            elif 'Calibre' in row:
                calibre = int(clean_currency(row.get('Calibre'))) if clean_currency(row.get('Calibre')) > 0 else None

            # Si Excel marca la fila como seleccionable y no pudimos inferir calibre,
            # dejamos un valor base para que la interfaz habilite selección por parte.
            if seleccionar_lamina and calibre is None:
                calibre = 12
            
            nuevo_bom = BomTemplate(
                modelo_elevador=str(row.get('Capacidad', '')),
                parte=str(row.get('Parte', '')),
                cantidad=int(round(clean_currency(row.get('cant.')))),
                material_referencia=material_text,
                peso_unitario=clean_currency(row.get('Peso')),
                costo_base_materia_prima=clean_currency(row.get('Valor Unit')),
                es_transformacion=is_lamina(material_text),
                calibre_lamina=calibre
            )
            db.add(nuevo_bom)
        print("[OK] Estructura BOM cargada (Precios limpios + calibres por parte).")

        db.commit()
        print("[OK] EXITO! Datos importados correctamente.")
    except Exception as e:
        db.rollback()
        print(f"[ERROR] Error critico: {e}")
    finally:
        db.close()

if __name__ == "__main__":
    importar_datos()
