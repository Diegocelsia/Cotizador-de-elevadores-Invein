from functools import lru_cache
from pathlib import Path
from typing import Dict, Optional, Tuple

from openpyxl import load_workbook

LAMINA_XLSX_PATH = Path(__file__).parent / "data" / "Lamina.xlsx"


@lru_cache(maxsize=1)
def _load_lamina_names(file_path: str = str(LAMINA_XLSX_PATH)) -> Dict[Tuple[str, int], str]:
    path = Path(file_path)
    if not path.exists():
        return {}

    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    index = {name: idx for idx, name in enumerate(header)}

    if "Material" not in index or "Tipo" not in index:
        return {}

    out: Dict[Tuple[str, int], str] = {}
    for row in rows[1:]:
        material_name = row[index["Material"]] if index["Material"] < len(row) else None
        tipo = row[index["Tipo"]] if index["Tipo"] < len(row) else None

        if not material_name or not tipo:
            continue

        material_text = str(material_name).strip()
        tipo_text = str(tipo).strip().upper()

        calibre = _extract_calibre(material_text)
        if calibre is None:
            continue

        out[(tipo_text, calibre)] = material_text

    return out


def _extract_calibre(text: str) -> Optional[int]:
    txt = str(text or "").lower().replace("á", "a")
    import re

    match = re.search(r"cal\.?\s*(\d{1,2})", txt)
    if not match:
        return None
    return int(match.group(1))


def get_lamina_name(tipo: str, calibre: int) -> str:
    key = (str(tipo or "").strip().upper(), int(calibre))
    names = _load_lamina_names()
    if key in names:
        return names[key]
    return f"Lam cal. {int(calibre)}"
