from functools import lru_cache
from pathlib import Path
from typing import Dict, Optional, Tuple

from openpyxl import load_workbook

MATERIAL_XLSX_PATH = Path(__file__).parent / "data" / "Material.xlsx"
LAMINA_XLSX_PATH = Path(__file__).parent / "data" / "Lamina.xlsx"


def _norm(value: str) -> str:
    txt = str(value or "").strip().lower()
    repl = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
    }
    for a, b in repl.items():
        txt = txt.replace(a, b)
    return " ".join(txt.split())


def _to_float(value) -> Optional[float]:
    if value in (None, "", "-"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except Exception:
        return None


@lru_cache(maxsize=1)
def _load_material_factors(file_path: str = str(MATERIAL_XLSX_PATH)) -> Dict[Tuple[str, str, str], float]:
    path = Path(file_path)
    if not path.exists():
        return {}

    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {name: i for i, name in enumerate(headers)}

    required = ["Capacidad", "Parte", "Material", "Factor"]
    for req in required:
        if req not in idx:
            return {}

    out: Dict[Tuple[str, str, str], float] = {}
    for row in rows[1:]:
        model = row[idx["Capacidad"]] if idx["Capacidad"] < len(row) else None
        part = row[idx["Parte"]] if idx["Parte"] < len(row) else None
        material = row[idx["Material"]] if idx["Material"] < len(row) else None
        factor_raw = row[idx["Factor"]] if idx["Factor"] < len(row) else None

        if not model or not part or not material:
            continue

        factor = _to_float(factor_raw)
        if factor is None:
            continue

        key = (_norm(model), _norm(part), _norm(material))
        out[key] = factor

    return out


@lru_cache(maxsize=1)
def _load_lamina_factors(file_path: str = str(LAMINA_XLSX_PATH)) -> Dict[Tuple[str, int], float]:
    path = Path(file_path)
    if not path.exists():
        return {}

    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {name: i for i, name in enumerate(headers)}

    if "Tipo" not in idx or "Material" not in idx or "Factor" not in idx:
        return {}

    import re

    out: Dict[Tuple[str, int], float] = {}
    for row in rows[1:]:
        tipo = row[idx["Tipo"]] if idx["Tipo"] < len(row) else None
        material_name = row[idx["Material"]] if idx["Material"] < len(row) else None
        factor_raw = row[idx["Factor"]] if idx["Factor"] < len(row) else None

        if not tipo or not material_name:
            continue

        mat_text = _norm(material_name)
        m = re.search(r"cal\.?\s*(\d{1,2})", mat_text)
        if not m:
            continue

        calibre = int(m.group(1))
        factor = _to_float(factor_raw)
        if factor is None:
            continue

        out[(_norm(tipo).upper(), calibre)] = factor

    return out


def get_factor_ganancia(
    modelo: str,
    parte: str,
    material_referencia: str,
    material_lamina_tipo: Optional[str] = None,
    calibre_lamina: Optional[int] = None,
) -> Tuple[Optional[float], str]:
    key = (_norm(modelo), _norm(parte), _norm(material_referencia))
    material_factors = _load_material_factors()
    factor = material_factors.get(key)

    if factor is not None and factor > 0:
        return factor, "Material.xlsx"

    # Para laminas genericas (factor 0 en Material.xlsx), usar factor por tipo/calibre.
    if material_lamina_tipo and calibre_lamina is not None:
        lamina_factors = _load_lamina_factors()
        lk = (_norm(material_lamina_tipo).upper(), int(calibre_lamina))
        lf = lamina_factors.get(lk)
        if lf is not None and lf > 0:
            return lf, "Lamina.xlsx"

    return None, "regla_default"
